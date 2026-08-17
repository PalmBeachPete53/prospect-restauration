"""Etape 5 : met le lot en forme dans un classeur, une feuille par niche.

Entree : lot_NN.csv     Sortie : lot_NN.xlsx (5 feuilles de 7 prospects)

Usage : python export_xlsx.py [lot_01.csv]
"""
import csv, sys, os
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config_prospect import NICHES

ARDOISE = '1F2937'      # en-tete
ENCRE = '111827'        # texte principal
GRIS = '6B7280'         # texte secondaire
ZEBRE = 'F9FAFB'        # ligne paire
BORDURE = 'E5E7EB'
ROUGE = 'B91C1C'        # alerte (pas de HTTPS, non responsive)
VERT = '047857'
# palette du score de laideur : plus c'est haut, plus c'est chaud
PALIERS = [(50, 'FEE2E2', '991B1B'), (35, 'FEF3C7', '92400E'), (0, 'F3F4F6', '374151')]

COLONNES = [
    ('Site web',        34),
    ('Nom',             26),
    ('Zone de recherche', 15),
    ('Commune réelle',  24),
    ('Dép.',             6),
    ('Laideur',          9),
    ('Ancien.',          9),
    ('Responsive',      11),
    ('HTTPS',            8),
    ('Techno',          20),
    ('Ce qui cloche',   62),
]

fin = Side(style='thin', color=BORDURE)
CADRE = Border(left=fin, right=fin, top=fin, bottom=fin)


def palier(score):
    for seuil, fond, texte in PALIERS:
        if score >= seuil:
            return fond, texte
    return PALIERS[-1][1], PALIERS[-1][2]


def ecrire_feuille(ws, libelle, lignes):
    ws.sheet_view.showGridLines = False

    # bandeau de titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLONNES))
    titre = ws.cell(row=1, column=1, value=f'{libelle}  ·  {len(lignes)} prospects')
    titre.font = Font(name='Calibri', size=15, bold=True, color='FFFFFF')
    titre.fill = PatternFill('solid', fgColor=ARDOISE)
    titre.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 34

    sous = ws.cell(row=2, column=1,
                   value='Classés du site le plus indéfendable au moins pire. '
                         '« Laideur » = score de design daté ; « Ancien. » = signaux d\'abandon. '
                         '« Zone » est le centre de recherche, pas l\'adresse : la commune réelle '
                         'n\'est indiquée que lorsque le site l\'affiche.')
    sous.font = Font(name='Calibri', size=9, italic=True, color=GRIS)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLONNES))
    ws.row_dimensions[2].height = 20

    # en-tetes
    for i, (nom, largeur) in enumerate(COLONNES, 1):
        c = ws.cell(row=3, column=i, value=nom)
        c.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=ARDOISE)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = CADRE
        ws.column_dimensions[get_column_letter(i)].width = largeur
    ws.row_dimensions[3].height = 26

    for n, r in enumerate(lignes):
        ligne = 4 + n
        zebre = PatternFill('solid', fgColor=ZEBRE) if n % 2 else None
        valeurs = [
            r['domaine'], r['nom'], r['ville'], r.get('commune', ''), r['dept'],
            int(r['design']), int(r['anciennete']),
            'non' if r['non_responsif'] == 'oui' else 'oui',
            'non' if r['sans_https'] == 'oui' else 'oui',
            r['techno'], r['signaux_design'],
        ]
        for i, v in enumerate(valeurs, 1):
            c = ws.cell(row=ligne, column=i, value=v)
            c.border = CADRE
            c.font = Font(name='Calibri', size=10, color=ENCRE)
            c.alignment = Alignment(vertical='center')
            if zebre:
                c.fill = zebre

        # lien cliquable
        lien = ws.cell(row=ligne, column=1)
        lien.hyperlink = f"http://{r['domaine']}"
        lien.font = Font(name='Calibri', size=10, color='1D4ED8', underline='single')

        # score de laideur mis en avant
        fond, texte = palier(int(r['design']))
        sc = ws.cell(row=ligne, column=6)
        sc.fill = PatternFill('solid', fgColor=fond)
        sc.font = Font(name='Calibri', size=11, bold=True, color=texte)
        sc.alignment = Alignment(horizontal='center', vertical='center')

        anc = ws.cell(row=ligne, column=7)
        anc.alignment = Alignment(horizontal='center', vertical='center')

        # non responsive / pas de HTTPS : c'est l'argument de vente
        for col in (8, 9):
            c = ws.cell(row=ligne, column=col)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.font = Font(name='Calibri', size=10, bold=c.value == 'non',
                          color=ROUGE if c.value == 'non' else VERT)

        # zone de recherche : ce n'est pas l'adresse, on la met en retrait
        zone = ws.cell(row=ligne, column=3)
        zone.alignment = Alignment(horizontal='left', vertical='center')
        zone.font = Font(name='Calibri', size=9, color=GRIS)
        # commune reelle : vide tant qu'elle n'a pas ete lue sur le site
        com = ws.cell(row=ligne, column=4)
        com.alignment = Alignment(horizontal='left', vertical='center')
        com.font = Font(name='Calibri', size=10, bold=bool(com.value), color=ENCRE)

        ws.cell(row=ligne, column=5).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=ligne, column=10).font = Font(name='Calibri', size=9, color=GRIS)
        ws.cell(row=ligne, column=11).font = Font(name='Calibri', size=9, color=GRIS)
        ws.cell(row=ligne, column=11).alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[ligne].height = 30

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(len(COLONNES))}{3 + len(lignes)}'


def main():
    source = sys.argv[1] if len(sys.argv) > 1 else 'lot_01.csv'
    with open(source, encoding='utf-8') as f:
        lignes = list(csv.DictReader(f, delimiter=';'))

    par_niche = OrderedDict((NICHES[k][0], []) for k in NICHES)
    for r in lignes:
        par_niche.setdefault(r['niche'], []).append(r)

    wb = Workbook()
    wb.remove(wb.active)
    for libelle, rows in par_niche.items():
        rows.sort(key=lambda r: (-int(r['design']), -int(r['anciennete'])))
        ws = wb.create_sheet(title=libelle[:31])
        ecrire_feuille(ws, libelle, rows)

    sortie = os.path.splitext(source)[0] + '.xlsx'
    wb.save(sortie)
    print(f'{sortie} : {len(wb.sheetnames)} feuilles, {len(lignes)} prospects')
    for libelle, rows in par_niche.items():
        scores = [int(r['design']) for r in rows]
        etendue = f'{min(scores)}-{max(scores)}' if scores else '-'
        print(f'  {libelle:<24} {len(rows)} prospects  laideur {etendue}')


if __name__ == '__main__':
    main()
